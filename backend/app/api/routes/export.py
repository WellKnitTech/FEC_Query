from fastapi import APIRouter, HTTPException, Query, Body
from fastapi.responses import StreamingResponse, Response
from typing import Optional, List, Dict
from pydantic import BaseModel
from app.services.fec_client import FECClient
from app.services.report_generator import ReportGenerator
from app.services.analysis import AnalysisService
import logging
import io
import csv
from io import StringIO, BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

router = APIRouter()


class RaceExportRequest(BaseModel):
    candidate_ids: List[str]
    office: str
    state: str
    district: Optional[str] = None
    year: Optional[int] = None
    format: str = "pdf"  # pdf, docx, md


def get_fec_client():
    """Get FEC client instance"""
    from app.services.container import get_service_container
    try:
        container = get_service_container()
        return container.get_fec_client()
    except ValueError as e:
        logger.error(f"FEC API key not configured: {e}")
        raise HTTPException(
            status_code=500,
            detail="FEC API key not configured. Please set FEC_API_KEY in your .env file."
        )


def get_report_generator():
    """Get report generator instance"""
    return ReportGenerator(get_fec_client())

def get_analysis_service():
    """Get analysis service instance"""
    return AnalysisService(get_fec_client())


@router.get("/candidate/{candidate_id}")
async def export_candidate(
    candidate_id: str,
    format: str = Query("pdf", regex="^(pdf|docx|md|csv|excel)$", description="Export format"),
    cycle: Optional[int] = Query(None, description="Election cycle")
):
    """Export candidate report in specified format"""
    try:
        report_generator = get_report_generator()
        
        # Collect all candidate data
        data = await report_generator.collect_candidate_data(candidate_id, cycle=cycle)
        
        if not data.get('candidate'):
            raise HTTPException(status_code=404, detail="Candidate not found")
        
        # Generate report based on format
        if format == "pdf":
            buffer = await report_generator.generate_pdf_report(data, is_race=False)
            filename = f"candidate_{candidate_id}_report.pdf"
            media_type = "application/pdf"
            return StreamingResponse(
                buffer,
                media_type=media_type,
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        elif format == "docx":
            buffer = await report_generator.generate_docx_report(data, is_race=False)
            filename = f"candidate_{candidate_id}_report.docx"
            media_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            return StreamingResponse(
                buffer,
                media_type=media_type,
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        elif format == "csv":
            buffer = await report_generator.generate_csv_export(data, is_race=False)
            filename = f"candidate_{candidate_id}_report.csv"
            media_type = "text/csv"
            return StreamingResponse(
                buffer,
                media_type=media_type,
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        elif format == "excel":
            buffer = await report_generator.generate_excel_export(data, is_race=False)
            filename = f"candidate_{candidate_id}_report.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            return StreamingResponse(
                buffer,
                media_type=media_type,
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        else:  # md
            markdown = await report_generator.generate_markdown_report(data, is_race=False)
            buffer = io.BytesIO(markdown.encode('utf-8'))
            filename = f"candidate_{candidate_id}_report.md"
            media_type = "text/markdown"
            return StreamingResponse(
                buffer,
                media_type=media_type,
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting candidate {candidate_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to export candidate report: {str(e)}")


@router.post("/race")
async def export_race(request: RaceExportRequest):
    """Export race report in specified format"""
    try:
        if request.format not in ["pdf", "docx", "md", "csv", "excel"]:
            raise HTTPException(status_code=400, detail="Invalid format. Must be pdf, docx, md, csv, or excel")
        
        if not request.candidate_ids:
            raise HTTPException(status_code=400, detail="At least one candidate ID is required")
        
        report_generator = get_report_generator()
        
        # Collect race data
        data = await report_generator.collect_race_data(
            candidate_ids=request.candidate_ids,
            office=request.office,
            state=request.state,
            district=request.district,
            year=request.year
        )
        
        # Generate report based on format
        if request.format == "pdf":
            buffer = await report_generator.generate_pdf_report(data, is_race=True)
            filename = f"race_{request.office}_{request.state}_report.pdf"
            if request.district:
                filename = f"race_{request.office}_{request.state}_{request.district}_report.pdf"
            media_type = "application/pdf"
            return StreamingResponse(
                buffer,
                media_type=media_type,
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        elif request.format == "docx":
            buffer = await report_generator.generate_docx_report(data, is_race=True)
            filename = f"race_{request.office}_{request.state}_report.docx"
            if request.district:
                filename = f"race_{request.office}_{request.state}_{request.district}_report.docx"
            media_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            return StreamingResponse(
                buffer,
                media_type=media_type,
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        elif request.format == "csv":
            buffer = await report_generator.generate_csv_export(data, is_race=True)
            filename = f"race_{request.office}_{request.state}_report.csv"
            if request.district:
                filename = f"race_{request.office}_{request.state}_{request.district}_report.csv"
            media_type = "text/csv"
            return StreamingResponse(
                buffer,
                media_type=media_type,
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        elif request.format == "excel":
            buffer = await report_generator.generate_excel_export(data, is_race=True)
            filename = f"race_{request.office}_{request.state}_report.xlsx"
            if request.district:
                filename = f"race_{request.office}_{request.state}_{request.district}_report.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            return StreamingResponse(
                buffer,
                media_type=media_type,
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        else:  # md
            markdown = await report_generator.generate_markdown_report(data, is_race=True)
            buffer = io.BytesIO(markdown.encode('utf-8'))
            filename = f"race_{request.office}_{request.state}_report.md"
            if request.district:
                filename = f"race_{request.office}_{request.state}_{request.district}_report.md"
            media_type = "text/markdown"
            return StreamingResponse(
                buffer,
                media_type=media_type,
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting race report: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to export race report: {str(e)}")


@router.get("/contributions/csv")
async def export_contributions_csv(
    candidate_id: Optional[str] = Query(None, description="Candidate ID"),
    committee_id: Optional[str] = Query(None, description="Committee ID"),
    contributor_name: Optional[str] = Query(None, description="Contributor name"),
    min_amount: Optional[float] = Query(None, description="Minimum amount"),
    max_amount: Optional[float] = Query(None, description="Maximum amount"),
    min_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    max_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    limit: int = Query(10000, ge=1, le=100000, description="Maximum results")
):
    """Export contributions as CSV"""
    try:
        fec_client = get_fec_client()
        contributions = await fec_client.get_contributions(
            candidate_id=candidate_id,
            committee_id=committee_id,
            contributor_name=contributor_name,
            min_amount=min_amount,
            max_amount=max_amount,
            min_date=min_date,
            max_date=max_date,
            limit=limit
        )
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Write headers
        writer.writerow([
            'Contribution ID', 'Candidate ID', 'Committee ID', 'Contributor Name',
            'Contributor City', 'Contributor State', 'Contributor Zip',
            'Contributor Employer', 'Contributor Occupation', 'Amount',
            'Date', 'Type', 'Receipt Type'
        ])
        
        # Write data
        for contrib in contributions:
            writer.writerow([
                contrib.get('sub_id') or contrib.get('contribution_id', ''),
                contrib.get('candidate_id', ''),
                contrib.get('committee_id', ''),
                contrib.get('contributor_name', ''),
                contrib.get('contributor_city', ''),
                contrib.get('contributor_state', ''),
                contrib.get('contributor_zip', ''),
                contrib.get('contributor_employer', ''),
                contrib.get('contributor_occupation', ''),
                contrib.get('contribution_amount', 0),
                contrib.get('contribution_receipt_date') or contrib.get('contribution_date', ''),
                contrib.get('contribution_type', ''),
                contrib.get('receipt_type', '')
            ])
        
        output.seek(0)
        buffer = BytesIO(output.getvalue().encode('utf-8'))
        filename = "contributions_export.csv"
        
        return StreamingResponse(
            buffer,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Error exporting contributions CSV: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to export contributions: {str(e)}")


@router.get("/contributions/excel")
async def export_contributions_excel(
    candidate_id: Optional[str] = Query(None, description="Candidate ID"),
    committee_id: Optional[str] = Query(None, description="Committee ID"),
    contributor_name: Optional[str] = Query(None, description="Contributor name"),
    min_amount: Optional[float] = Query(None, description="Minimum amount"),
    max_amount: Optional[float] = Query(None, description="Maximum amount"),
    min_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    max_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    limit: int = Query(10000, ge=1, le=100000, description="Maximum results")
):
    """Export contributions as Excel"""
    try:
        fec_client = get_fec_client()
        contributions = await fec_client.get_contributions(
            candidate_id=candidate_id,
            committee_id=committee_id,
            contributor_name=contributor_name,
            min_amount=min_amount,
            max_amount=max_amount,
            min_date=min_date,
            max_date=max_date,
            limit=limit
        )
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Contributions"
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        headers = [
            'Contribution ID', 'Candidate ID', 'Committee ID', 'Contributor Name',
            'Contributor City', 'Contributor State', 'Contributor Zip',
            'Contributor Employer', 'Contributor Occupation', 'Amount',
            'Date', 'Type', 'Receipt Type'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data
        for row_idx, contrib in enumerate(contributions, 2):
            ws.cell(row=row_idx, column=1, value=contrib.get('sub_id') or contrib.get('contribution_id', ''))
            ws.cell(row=row_idx, column=2, value=contrib.get('candidate_id', ''))
            ws.cell(row=row_idx, column=3, value=contrib.get('committee_id', ''))
            ws.cell(row=row_idx, column=4, value=contrib.get('contributor_name', ''))
            ws.cell(row=row_idx, column=5, value=contrib.get('contributor_city', ''))
            ws.cell(row=row_idx, column=6, value=contrib.get('contributor_state', ''))
            ws.cell(row=row_idx, column=7, value=contrib.get('contributor_zip', ''))
            ws.cell(row=row_idx, column=8, value=contrib.get('contributor_employer', ''))
            ws.cell(row=row_idx, column=9, value=contrib.get('contributor_occupation', ''))
            ws.cell(row=row_idx, column=10, value=contrib.get('contribution_amount', 0))
            ws.cell(row=row_idx, column=11, value=contrib.get('contribution_receipt_date') or contrib.get('contribution_date', ''))
            ws.cell(row=row_idx, column=12, value=contrib.get('contribution_type', ''))
            ws.cell(row=row_idx, column=13, value=contrib.get('receipt_type', ''))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception as e:
                    logger.debug(f"Error reading cell value for column width calculation: {e}")
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = "contributions_export.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Error exporting contributions Excel: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to export contributions: {str(e)}")


@router.get("/out-of-state-contributions/csv")
async def export_out_of_state_contributions_csv(
    candidate_id: str = Query(..., description="Candidate ID"),
    min_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    max_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    cycle: Optional[int] = Query(None, description="Election cycle"),
    limit: int = Query(10000, ge=1, le=100000, description="Maximum results")
):
    """Export out-of-state contributions as CSV"""
    try:
        analysis_service = get_analysis_service()
        contributions = await analysis_service.get_out_of_state_contributions(
            candidate_id=candidate_id,
            min_date=min_date,
            max_date=max_date,
            cycle=cycle,
            limit=limit
        )
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Write headers
        writer.writerow([
            'Contribution ID', 'Candidate ID', 'Committee ID', 'Contributor Name',
            'Contributor City', 'Contributor State', 'Contributor Zip',
            'Contributor Employer', 'Contributor Occupation', 'Amount',
            'Date', 'Type', 'Receipt Type'
        ])
        
        # Write data
        for contrib in contributions:
            writer.writerow([
                contrib.get('contribution_id') or contrib.get('sub_id', ''),
                contrib.get('candidate_id', ''),
                contrib.get('committee_id', ''),
                contrib.get('contributor_name', ''),
                contrib.get('contributor_city', ''),
                contrib.get('contributor_state', ''),
                contrib.get('contributor_zip', ''),
                contrib.get('contributor_employer', ''),
                contrib.get('contributor_occupation', ''),
                contrib.get('contribution_amount', 0),
                contrib.get('contribution_date') or contrib.get('contribution_receipt_date', ''),
                contrib.get('contribution_type', ''),
                contrib.get('receipt_type', '')
            ])
        
        output.seek(0)
        buffer = BytesIO(output.getvalue().encode('utf-8'))
        filename = f"out_of_state_contributions_{candidate_id}.csv"
        
        return StreamingResponse(
            buffer,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Error exporting out-of-state contributions CSV: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to export out-of-state contributions: {str(e)}")


@router.get("/out-of-state-donors/csv")
async def export_out_of_state_donors_csv(
    candidate_id: str = Query(..., description="Candidate ID"),
    min_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    max_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    cycle: Optional[int] = Query(None, description="Election cycle"),
    limit: int = Query(1000, ge=1, le=10000, description="Maximum results")
):
    """Export aggregated out-of-state donors as CSV"""
    try:
        analysis_service = get_analysis_service()
        donors = await analysis_service.get_aggregated_out_of_state_donors(
            candidate_id=candidate_id,
            min_date=min_date,
            max_date=max_date,
            cycle=cycle,
            limit=limit
        )
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Write headers
        writer.writerow([
            'Donor Key', 'Canonical Name', 'All Name Variations', 'Total Amount',
            'Contribution Count', 'State', 'City', 'Employer', 'Occupation',
            'First Contribution Date', 'Last Contribution Date', 'Match Confidence',
            'Contribution IDs'
        ])
        
        # Write data
        for donor in donors:
            all_names = donor.get('all_names', [])
            all_names_str = '; '.join(all_names) if all_names else ''
            contribution_ids = donor.get('contribution_ids', [])
            contribution_ids_str = '; '.join(contribution_ids) if contribution_ids else ''
            
            writer.writerow([
                donor.get('donor_key', ''),
                donor.get('canonical_name', ''),
                all_names_str,
                donor.get('total_amount', 0),
                donor.get('contribution_count', 0),
                donor.get('canonical_state', ''),
                donor.get('canonical_city', ''),
                donor.get('canonical_employer', ''),
                donor.get('canonical_occupation', ''),
                donor.get('first_contribution_date', ''),
                donor.get('last_contribution_date', ''),
                donor.get('match_confidence', 0),
                contribution_ids_str
            ])
        
        output.seek(0)
        buffer = BytesIO(output.getvalue().encode('utf-8'))
        filename = f"out_of_state_donors_{candidate_id}.csv"
        
        return StreamingResponse(
            buffer,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Error exporting out-of-state donors CSV: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to export out-of-state donors: {str(e)}")

